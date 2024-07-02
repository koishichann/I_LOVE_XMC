'''
This model is used to generate keyphrase with raw documents
functions: data processing, training, predicting
'''
import os
from transformers import (PegasusTokenizerFast, PegasusTokenizer, BartForConditionalGeneration,
                          PegasusForConditionalGeneration, T5Tokenizer, T5ForConditionalGeneration,
                          AutoTokenizer, AutoModel, BartTokenizer, MBartForConditionalGeneration, MBart50TokenizerFast,
                          GPT2Tokenizer, GPT2Model, BatchEncoding, BartTokenizerFast
                          )
import torch
from torch.utils.data import DataLoader
import wandb
import pytorch_lightning as pl
from tqdm import tqdm
from pytorch_lightning.callbacks import EarlyStopping, ModelCheckpoint, LearningRateMonitor
from ..utils.utils import load_map, read_index, read_text, transfer_indexs_to_labels
from ..utils import utils
from opendelta import AutoDeltaModel, AutoDeltaConfig, PrefixModel
from bigmodelvis import Visualization
import openpyxl
import re

device = 'cuda' if torch.cuda.is_available() else 'cpu'


def valid_xml_char_ordinal(c):
    codepoint = ord(c)
    # conditions ordered by presumed frequency
    return (
            0x20 <= codepoint <= 0xD7FF or
            codepoint in (0x9, 0xA, 0xD) or
            0xE000 <= codepoint <= 0xFFFD or
            0x10000 <= codepoint <= 0x10FFFF
    )


def data_clean(text):
    # 清洗excel中的非法字符，都是不常见的不可显示字符，例如退格，响铃等
    ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
    text = ILLEGAL_CHARACTERS_RE.sub(r'', str(text))
    text = ''.join(c for c in text if valid_xml_char_ordinal(c))
    return text


class MyData(torch.utils.data.Dataset):
    def __init__(self, encoding, labels):
        self.ids = encoding['input_ids']
        self.mask = encoding['attention_mask']
        self.labels = labels['input_ids']

    def __getitem__(self, idx):
        item = {}
        item['input_ids'] = torch.tensor(self.ids[idx]).to(device)
        item['attention_mask'] = torch.tensor(self.mask[idx]).to(device)
        item['labels'] = torch.tensor(self.labels[idx]).to(device)
        # item={'input_ids': torch.tensor(val[idx]).to(device) for key, val in self.encoding.items()}
        # item['labels'] = torch.tensor(self.labels['input_ids'][idx]).to(device)
        return item

    def __len__(self):
        return len(self.labels)  # len(self.labels)


class KG_Model(pl.LightningModule):
    def __init__(self, args) -> None:
        super().__init__()
        self.model_name = args.kg_model_name
        self.batch_size = args.kg_batch_size
        self.lr = args.kg_lr
        # self.epoch = args.kg_epoch
        self.type = args.kg_type
        self.datadir = args.datadir
        # self.trn = args.kg_trn_data
        # self.tst = args.kg_tst_data
        self.top_k = args.top_k
        self.top_p = None
        self.prefix_token_num = args.prefix_token_num
        if 'two' in self.type:
            self.two = True
            self.current_model = 'present'
        else:
            self.two = False
            self.current_model = 'all'
        self.train_dataset = None
        self.val_dataset = None
        if 'shuffle' in self.type:
            self.shuffle = True
        else:
            self.shuffle = False

        if 'kpdrop_na' in self.type:
            self.kpdrop = 'na'
        elif 'kpdrop_nr' in self.type:
            self.kpdrop = 'nr'
        elif 'kpdrop_a' in self.type:
            self.kpdrop = 'a'
        elif 'kpdrop_r' in self.type:
            self.kpdrop = 'r'
        else:
            self.kpdrop = None
        self.kpdrop_rate = args.kpdrop_rate

        if 'kpinsert_na' in self.type:
            self.kpinsert = 'na'
        elif 'kpinsert_nr' in self.type:
            self.kpinsert = 'nr'
        elif 'kpinsert_a' in self.type:
            self.kpinsert = 'a'
        elif 'kpinsert_r' in self.type:
            self.kpinsert = 'r'
        else:
            self.kpinsert = None
        self.kpinsert_rate = args.kpinsert_rate

        if args.top_p:
            self.top_p = args.top_p
        if args.max_len:
            self.max_len = args.max_len

        self.match = args.match
        self.stem_delta = args.stem_delta
        self.stem_lambda = args.stem_lambda
        self.stem_model = args.stem_model
        # self.save_dir = args.kg_savedir
        self.curr_avg_loss = 0.0
        self.train_present_ratio, self.train_absent_ratio, self.val_present_ratio, self.val_absent_ratio = 0.0, 0.0, 0.0, 0.0
        self.train_label_word_num_map, self.val_label_word_num_map = {}, {}
        self.wb = None
        if 'pega' in self.type:
            self.model = PegasusForConditionalGeneration.from_pretrained(self.model_name).to(device)
            self.tokenizer = PegasusTokenizer.from_pretrained(self.model_name, model_max_length=self.max_len)
        elif 'mbart' in self.model_name:
            self.model = MBartForConditionalGeneration.from_pretrained(self.model_name).to(device)
            self.tokenizer = MBart50TokenizerFast.from_pretrained(self.model_name, model_max_length=self.max_len)
        elif 'bart' in self.type:
            if 'two' in self.type:
                self.present_model = BartForConditionalGeneration.from_pretrained(self.model_name).to(device)
                self.absent_model = BartForConditionalGeneration.from_pretrained(self.model_name).to(device)
                if 'prefix' in self.type:
                    prefix_bart_model_present = PrefixModel(backbone_model=self.present_model,
                                                            modified_modules=['encoder.layers.0.self_attn',
                                                                              'encoder.layers.1.self_attn',
                                                                              'encoder.layers.2.self_attn',
                                                                              'encoder.layers.3.self_attn',
                                                                              'encoder.layers.4.self_attn',
                                                                              'encoder.layers.5.self_attn'],
                                                            reparameterize=False,
                                                            prefix_token_num=self.prefix_token_num).to(
                        device)
                    prefix_bart_model_present.freeze_module(exclude=['deltas', 'layernorm_embedding'],
                                                            set_state_dict=True)
                    prefix_bart_model_present.log()
                    prefix_bart_model_absent = PrefixModel(backbone_model=self.absent_model,
                                                           modified_modules=['encoder.layers.0.self_attn',
                                                                             'encoder.layers.1.self_attn',
                                                                             'encoder.layers.2.self_attn',
                                                                             'encoder.layers.3.self_attn',
                                                                             'encoder.layers.4.self_attn',
                                                                             'encoder.layers.5.self_attn'],
                                                           reparameterize=False,
                                                           prefix_token_num=self.prefix_token_num).to(
                        device)
                    prefix_bart_model_absent.freeze_module(exclude=['deltas', 'layernorm_embedding'],
                                                           set_state_dict=True)
                    prefix_bart_model_absent.log()
            else:
                self.model = BartForConditionalGeneration.from_pretrained(self.model_name).to(device)
                if 'prefix' in self.type:
                    prefix_bart_model = PrefixModel(backbone_model=self.model,
                                                    modified_modules=['encoder.layers.0.self_attn',
                                                                      'encoder.layers.1.self_attn',
                                                                      'encoder.layers.2.self_attn',
                                                                      'encoder.layers.3.self_attn',
                                                                      'encoder.layers.4.self_attn',
                                                                      'encoder.layers.5.self_attn'],
                                                    reparameterize=False, prefix_token_num=self.prefix_token_num).to(
                        device)
                    prefix_bart_model.freeze_module(exclude=['deltas', 'layernorm_embedding'], set_state_dict=True)
                    prefix_bart_model.log()
                # Visualization(self.model).structure_graph()
            self.tokenizer = BartTokenizerFast.from_pretrained(self.model_name, model_max_length=self.max_len)
        elif 't5' in self.type:
            self.model = T5ForConditionalGeneration.from_pretrained(self.model_name).to(device)
            self.tokenizer = T5Tokenizer.from_pretrained(self.model_name, model_max_length=self.max_len)
        elif 'gpt2' in self.model_name:
            self.model = GPT2Model.from_pretrained(self.model_name).to(device)
            self.tokenizer = GPT2Tokenizer.from_pretrained(self.model_name, model_max_length=self.max_len)
        else:
            print('未识别model')
            self.model = AutoModel.from_pretrained(self.model_name)
            self.tokenizer = AutoTokenizer.from_pretrained(self.model_name)

        self.save_hyperparameters()
        # forward part

    def forward(self, encoder_input_ids, labels):
        if self.two:
            if self.current_model == 'present':
                return self.present_model(input_ids=encoder_input_ids, labels=labels)
            elif self.current_model == 'absent':
                return self.absent_model(input_ids=encoder_input_ids, labels=labels)
        return self.model(input_ids=encoder_input_ids, labels=labels)

    def training_step(self, batch, batch_idx):
        encoder_input_ids, encoder_attention_mask, labels = torch.stack([i['input_ids'] for i in batch]), torch.stack(
            [i['attention_mask'] for i in batch]), torch.stack([i['labels'] for i in batch])
        res = self(encoder_input_ids, labels)
        cur_lr = self.trainer.optimizers[0].param_groups[0]['lr']
        global_step = self.trainer.global_step
        # #手动优化scheduler
        sch = self.lr_schedulers()
        loss = res.loss
        self.curr_avg_loss += loss
        if (global_step + 1) % 50 == 0:
            wandb.log(
                {self.current_model + "_loss": self.curr_avg_loss / 50,
                 self.current_model + "_global_step": global_step})
            wandb.log({self.current_model + "_learning_rate": cur_lr, self.current_model + "_global_step": global_step})
            wandb.log({self.current_model + "_train_epoch": self.trainer.current_epoch,
                       self.current_model + "_global_step": global_step})
            self.curr_avg_loss = 0.0
        if (batch_idx + 1) % 5 == 0:
            sch.step()
        self.log('lr', cur_lr, prog_bar=True, on_step=True)
        return loss

    def validation_step(self, batch, batch_idx):
        encoder_input_ids, encoder_attention_mask, labels = torch.stack([i['input_ids'] for i in batch]), torch.stack(
            [i['attention_mask'] for i in batch]), torch.stack([i['labels'] for i in batch])
        res = self(encoder_input_ids=encoder_input_ids, labels=labels)
        self.log(self.current_model + '_val_loss', res.loss, prog_bar=True, batch_size=self.batch_size)
        return res.loss

    def configure_optimizers(self):
        optimizer = torch.optim.Adam(self.parameters(), lr=self.lr)

        return {
            "optimizer": optimizer,
            "lr_scheduler": torch.optim.lr_scheduler.StepLR(optimizer=optimizer, step_size=1, gamma=0.9995),
            "interval": "step",
            "frequency": 1,
        }

    def generate(self, input_ids, attention_mask, max_length, num_beams):
        if self.top_p and self.top_k:
            if self.two:
                if self.current_model == 'present':
                    return self.present_model.generate(input_ids=input_ids, attention_mask=attention_mask,
                                                       max_length=max_length, num_beams=num_beams, top_p=self.top_p)
                elif self.current_model == 'absent':
                    return self.absent_model.generate(input_ids=input_ids, attention_mask=attention_mask,
                                                      max_length=max_length, num_beams=num_beams, top_p=self.top_p)
            return self.model.generate(input_ids=input_ids, attention_mask=attention_mask,
                                       max_length=max_length, num_beams=num_beams, top_p=self.top_p)
        else:
            if self.two:
                if self.current_model == 'present':
                    return self.present_model.generate(input_ids=input_ids, attention_mask=attention_mask,
                                                       max_length=max_length, num_beams=num_beams)
                elif self.current_model == 'absent':
                    return self.absent_model.generate(input_ids=input_ids, attention_mask=attention_mask,
                                                      max_length=max_length, num_beams=num_beams)
            return self.model.generate(input_ids=input_ids, attention_mask=attention_mask,
                                       max_length=max_length, num_beams=num_beams)

    def getTokenized(self, texts, printString="\nstart processing:", processName=""):
        data = {'input_ids': [], 'attention_mask': []}
        encodings = []
        print(printString + ' ' + processName)
        for text in texts:
            token = self.tokenizer(text, truncation=True, padding=True)
            data['input_ids'].append(token.data['input_ids'])
            data['attention_mask'].append(token.data['attention_mask'])
            encodings.append(token.encodings)
        result = BatchEncoding(data=data, encoding=encodings, n_sequences=1)
        return result

    def kind_dataloader(self, datadir, kind, type, epoch):
        prefix = "Summary: "
        texts = read_text(os.path.join(datadir, "X." + kind + '.' + type + "." + str(epoch) + ".txt"))
        labels = read_text(os.path.join(datadir, "Y_mapped." + kind + '.' + type + "." + str(epoch) + ".txt"))
        if 'wiki' in datadir:
            for id, label in enumerate(labels):
                labels[id] = label.replace('_', ' ')
        texts = list(map(lambda x: prefix + x, texts))

        '''
        在这里加入shuffle或者sort,改变train_labels
        '''

        print("\nstart processing: encoding " + kind + ' ' + type + " texts")
        encodings = self.tokenizer(texts, truncation=True, padding=True)
        print("\nstart processing: decoding " + kind + ' ' + type + " texts")
        decodings = self.tokenizer(labels, truncation=True, padding=True)

        dataset_tokenized = MyData(encodings, decodings)
        data = DataLoader(dataset_tokenized, batch_size=self.batch_size, collate_fn=lambda x: x, shuffle=True)
        return data

    def train_dataloader(self, datadir):
        prefix = "Summary: "
        # 获取text, 获取label index，映射出label text
        train_texts = read_text(os.path.join(datadir, "X.trn.txt"))
        train_indexes = read_index(os.path.join(datadir, "Y.trn.txt"))
        label_map = load_map(os.path.join(datadir, "output-items.txt"))
        train_labels_list = transfer_indexs_to_labels(label_map, train_indexes)  # list,需要转化成text

        if self.two:
            train_present_label_list, train_absent_label_list, train_present_indexes, train_absent_indexes, self.train_present_ratio, self.train_absent_ratio, self.train_label_word_num_map = utils.separate_present_absent_labels(
                train_indexes, train_labels_list, train_texts, self.match, self.stem_lambda, self.stem_delta,
                self.stem_model)
            self.store_label_word_num_map(self.train_label_word_num_map, 'train')
            if self.kpdrop is not None or self.kpinsert is not None:
                if self.kpdrop is not None:
                    train_present_label_list, train_absent_label_list, train_present_texts, train_absent_texts = utils.kpdrop(
                        present_labels=train_present_label_list, absent_labels=train_absent_label_list,
                        texts=train_texts, kpdrop_type=self.kpdrop, kpdrop_rate=self.kpdrop_rate)
                if self.kpinsert is not None:
                    train_present_label_list, train_absent_label_list, train_present_texts, train_absent_texts = utils.kpinsert(
                        present_labels=train_present_label_list, absent_labels=train_absent_label_list,
                        texts=train_texts, kpinsert_type=self.kpinsert, kpinsert_rate=self.kpinsert_rate)
            else:
                train_present_texts = train_texts
                train_absent_texts = train_texts.copy()
            if self.shuffle:
                utils.add_shuffle_examples(train_present_label_list, train_present_texts)
                utils.add_shuffle_examples(train_absent_label_list, train_absent_texts)
            train_present_labels, train_absent_labels = [], []
            for i in train_present_label_list:
                train_present_labels.append(", ".join(i))  # 是否加prefix
            for i in train_absent_label_list:
                train_absent_labels.append(", ".join(i))  # 是否加prefix
            train_present_texts = list(map(lambda x: prefix + x, train_present_texts))
            train_absent_texts = list(map(lambda x: prefix + x, train_absent_texts))

            '''
            在这里加入shuffle或者sort,改变train_labels
            '''

            print("\nstart processing: encoding train texts")
            present_encodings = self.tokenizer(train_present_texts, truncation=True, padding=True)
            absent_encodings = self.tokenizer(train_absent_texts, truncation=True, padding=True)
            print("\nstart processing: decoding train texts")
            present_decodings = self.tokenizer(train_present_labels, truncation=True, padding=True)
            absent_decodings = self.tokenizer(train_absent_labels, truncation=True, padding=True)

            present_dataset_tokenized = MyData(present_encodings, present_decodings)
            absent_dataset_tokenized = MyData(absent_encodings, absent_decodings)
            present_train_data = DataLoader(
                present_dataset_tokenized, batch_size=self.batch_size, collate_fn=lambda x: x, shuffle=True)
            absent_train_data = DataLoader(
                absent_dataset_tokenized, batch_size=self.batch_size, collate_fn=lambda x: x, shuffle=True)
            self.train_dataset = {'present': present_train_data, 'absent': absent_train_data}
        else:
            train_labels = []
            for i in train_labels_list:
                train_labels.append(", ".join(i))  # 是否加prefix
            train_texts = list(map(lambda x: prefix + x, train_texts))

            '''
            在这里加入shuffle或者sort,改变train_labels
            '''

            print("\nstart processing: encoding train texts")
            encodings = self.tokenizer(train_texts, truncation=True, padding=True)
            print("\nstart processing: decoding train texts")
            decodings = self.tokenizer(train_labels, truncation=True, padding=True)

            dataset_tokenized = MyData(encodings, decodings)
            train_data = DataLoader(
                dataset_tokenized, batch_size=self.batch_size, collate_fn=lambda x: x, shuffle=True)
            self.train_dataset = {'all': train_data}
        # create a dataloader for your training data here

    def val_dataloader(self, datadir):
        prefix = "Summary: "
        # 获取text, 获取label index，映射出label text
        val_texts = read_text(os.path.join(datadir, "X.tst.txt"))
        val_indexes = read_index(os.path.join(datadir, "Y.tst.txt"))
        # val_index = list(map(lambda x: x if len(x)<10 else x[0:10],val_index))
        label_map = load_map(os.path.join(datadir, "output-items.txt"))
        val_labels_list = transfer_indexs_to_labels(label_map, val_indexes)  # list,需要转化成text

        if self.two:
            val_present_label_list, val_absent_label_list, val_present_indexes, val_absent_indexes, self.val_present_ratio, self.val_absent_ratio, self.val_label_word_num_map = utils.separate_present_absent_labels(
                val_indexes, val_labels_list, val_texts, self.match, self.stem_lambda, self.stem_delta, self.stem_model)
            self.store_label_word_num_map(self.val_label_word_num_map, 'val')
            if self.kpdrop is not None or self.kpinsert is not None:
                if self.kpdrop is not None:
                    val_present_label_list, val_absent_label_list, val_present_texts, val_absent_texts = utils.kpdrop(
                        present_labels=val_present_label_list, absent_labels=val_absent_label_list,
                        texts=val_texts, kpdrop_type=self.kpdrop, kpdrop_rate=self.kpdrop_rate)
                if self.kpinsert is not None:
                    val_present_label_list, val_absent_label_list, val_present_texts, val_absent_texts = utils.kpinsert(
                        present_labels=val_present_label_list, absent_labels=val_absent_label_list,
                        texts=val_texts, kpinsert_type=self.kpinsert, kpinsert_rate=self.kpinsert_rate)
            else:
                val_present_texts = val_texts
                val_absent_texts = val_texts.copy()
            if self.shuffle:
                utils.add_shuffle_examples(val_present_label_list, val_present_texts)
                utils.add_shuffle_examples(val_absent_label_list, val_absent_texts)
            val_present_labels, val_absent_labels = [], []
            for i in val_present_label_list:
                val_present_labels.append(", ".join(i))  # 是否加prefix
            for i in val_absent_label_list:
                val_absent_labels.append(", ".join(i))  # 是否加prefix
            val_present_texts = list(map(lambda x: prefix + x, val_present_texts))
            val_absent_texts = list(map(lambda x: prefix + x, val_absent_texts))
            '''
            在这里加入shuffle或者sort,改变val_labels
            '''

            print("\nstart processing: encoding val texts")
            present_encodings = self.tokenizer(val_present_texts, truncation=True, padding=True)
            absent_encodings = self.tokenizer(val_absent_texts, truncation=True, padding=True)
            print("\nstart processing: decoding val texts")
            present_decodings = self.tokenizer(val_present_labels, truncation=True, padding=True)
            absent_decodings = self.tokenizer(val_absent_labels, truncation=True, padding=True)

            present_dataset_tokenized = MyData(present_encodings, present_decodings)
            absent_dataset_tokenized = MyData(absent_encodings, absent_decodings)
            present_val_data = DataLoader(
                present_dataset_tokenized, batch_size=self.batch_size, collate_fn=lambda x: x, shuffle=True)
            absent_val_data = DataLoader(
                absent_dataset_tokenized, batch_size=self.batch_size, collate_fn=lambda x: x, shuffle=True)
            self.val_dataset = {'present': present_val_data, 'absent': absent_val_data}
        else:
            val_labels = []
            for i in val_labels_list:
                val_labels.append(",".join(i))  # 是否加prefix
            val_texts = list(map(lambda x: prefix + x, val_texts))
            '''
            在这里加入shuffle或者sort,改变train_labels
            '''

            print("\nstart processing: encoding val texts")
            encodings = self.tokenizer(val_texts, truncation=True, padding=True)
            print("\nstart processing: decoding val texts")
            decodings = self.tokenizer(val_labels, truncation=True, padding=True)

            dataset_tokenized = MyData(encodings, decodings)
            val_data = DataLoader(
                dataset_tokenized, batch_size=self.batch_size, collate_fn=lambda x: x, shuffle=True)
            # create a dataloader for your training data here
            self.val_dataset = {'all': val_data}

    def load_kind_data(self, datadir, epoch):
        trn_data = self.kind_dataloader(datadir=datadir, kind=self.current_model, type='trn', epoch=epoch)
        tst_data = self.kind_dataloader(datadir=datadir, kind=self.current_model, type='tst', epoch=epoch)
        return trn_data, tst_data

    def load_data(self, datadir):
        self.train_dataloader(datadir=datadir)
        self.val_dataloader(datadir=datadir)

    def switch_present_and_absent(self):
        if self.two:
            if self.current_model == 'present':
                self.current_model = 'absent'
            elif self.current_model == 'absent':
                self.current_model = 'present'


def get_predict(documents, tokenizer, model, sheet, id):
    id_ = id
    inputs = tokenizer(documents, return_tensors='pt', padding=True, truncation=True).to(device)
    if model.two:
        present_summary_ids = model.generate(inputs['input_ids'], inputs['attention_mask'], max_length=128,
                                             num_beams=5).to(
            device)
        present_pre_result = tokenizer.batch_decode(present_summary_ids, skip_special_tokens=True,
                                                    clean_up_tokenization_spaces=True,
                                                    pad_to_multiple_of=2)
        for t in present_pre_result:
            if id > 100000:
                break
            sheet['B' + str(id + 2)] = data_clean(t)
            id = id + 1
        model.switch_present_and_absent()
        absent_summary_ids = model.generate(inputs['input_ids'], inputs['attention_mask'], max_length=128,
                                            num_beams=5).to(
            device)
        absent_pre_result = tokenizer.batch_decode(absent_summary_ids, skip_special_tokens=True,
                                                   clean_up_tokenization_spaces=True,
                                                   pad_to_multiple_of=2)
        for t in absent_pre_result:
            if id_ > 100000:
                break
            sheet['C' + str(id_ + 2)] = data_clean(t)
            id_ = id_ + 1
        model.switch_present_and_absent()
        pre_result = []
        for i, present_line in enumerate(present_pre_result):
            pre_result.append(present_line + ', ' + absent_pre_result[i])
    else:
        summary_ids = model.generate(inputs['input_ids'], inputs['attention_mask'], max_length=128, num_beams=5).to(
            device)
        pre_result = tokenizer.batch_decode(summary_ids, skip_special_tokens=True, clean_up_tokenization_spaces=True,
                                            pad_to_multiple_of=2)
    return pre_result, id


def kg_predict(model, args, type):
    '''
    model is the trained model of pretrained text2text model like BART.
    tokenizer is the tokenizer which is compaitible for model.
    src is the complete relateive dir of documents Y.tst.txt.
    outputdir is the output dir which is in the datadir.
    datasize is the batch_decode size.
    '''

    datadir = utils.BASE_DATASET_DIR
    src_dir = os.path.join(utils.BASE_DATASET_DIR, 'X.' + type + '.txt')
    output_dir = os.path.join(utils.KEYPHRASE_GENERATION_RECORDS_DIR, 'res', utils.KEYPHRASE_GENERATION_MODEL_NAME,
                              type + '_pred.txt')
    xlsx_dir = os.path.join(utils.KEYPHRASE_GENERATION_RECORDS_DIR, 'res', utils.KEYPHRASE_GENERATION_MODEL_NAME, type +
                            '_generated_keyphrases.xlsx')
    if os.path.exists(output_dir):
        return
    data_size = args.kg_batch_size
    print(f'src_dir: {src_dir}')
    print(f'output_dir: {output_dir}')
    res = []
    doc_list = []
    with open(src_dir, 'r+', encoding='utf-8', errors='ignore') as f:
        for i in f:
            doc_list.append(i)

    wb = openpyxl.Workbook()
    wb.save(filename=xlsx_dir)
    wb = openpyxl.load_workbook(xlsx_dir)
    wb.create_sheet('generated_keyphrases', 0)
    sheet = wb['generated_keyphrases']
    sheet['A1'] = data_clean('document')
    if model.two:
        sheet['B1'] = data_clean('present_words')
        sheet['C1'] = data_clean('absent_words')
    else:
        sheet['B1'] = data_clean('words')
    for id, i in enumerate(doc_list):
        if id > 100000:
            break
        sheet['A' + str(id + 2)] = data_clean(i)
    dataloader = DataLoader(doc_list, batch_size=data_size)
    tokenizer = model.tokenizer

    with open(output_dir, 'w+', encoding='utf-8', errors='ignore') as t:
        id = 0
        for i in tqdm(dataloader):  # range(len(data))
            tmp_result, id = get_predict(documents=i, tokenizer=tokenizer, model=model, sheet=sheet, id=id)
            for j in tmp_result:
                l_labels = []  # l_label 是str转 label的集合
                pre = j.replace("Summary: ", "").strip().split(", ")
                for k in range(len(pre)):
                    tmpstr = pre[k].strip(" ").strip("'").strip('"')
                    if tmpstr == '': continue
                    l_labels.append(tmpstr)
                res.append(l_labels)
                t.write(" || ".join(l_labels))
                t.write("\n")
        wb.save(xlsx_dir)


def kg_train(args):
    ###
    model_dir = os.path.join(utils.KEYPHRASE_GENERATION_RECORDS_DIR, 'res', utils.KEYPHRASE_GENERATION_MODEL_NAME,
                             'model_save')
    xlsx_dir = os.path.join(utils.KEYPHRASE_GENERATION_RECORDS_DIR, 'res', utils.KEYPHRASE_GENERATION_MODEL_NAME,
                            'intermediate_data.xlsx')
    # datadir = os.path.join(utils.KEYPHRASE_GENERATION_RECORDS_DIR, 'res', utils.KEYPHRASE_GENERATION_MODEL_NAME,
    #                       'preprocessed data')

    if os.path.exists(model_dir) and 'two' not in args.kg_type:
        return model_dir
    if os.path.exists(model_dir + '_present') and os.path.exists(model_dir + '_absent') and 'two' in args.kg_type:
        return model_dir

    # early_stopping = EarlyStopping(monitor='val_loss', patience=3, mode='min')

    checkpoint_callback = ModelCheckpoint(
        dirpath='./log/kg_check',
        filename='{epoch}-{val_loss:.2f}-{other_metric:.2f}'
    )
    lr_callback = LearningRateMonitor(logging_interval="step")
    model = KG_Model(args)
    # model.create_xlsx(xlsx_dir)
    # model.load_data(datadir=utils.BASE_DATASET_DIR)
    # model.save_xlsx(xlsx_dir)
    if model.two:
        present_early_stopping = EarlyStopping(monitor='present_val_loss', patience=3, mode='min')
        absent_early_stopping = EarlyStopping(monitor='absent_val_loss', patience=3, mode='min')
        datadir = os.path.join(utils.KEYPHRASE_GENERATION_RECORDS_DIR, 'res', utils.KEYPHRASE_GENERATION_MODEL_NAME,
                               'preprocessed data')

        for e in range(args.kg_epoch):
            torch.cuda.empty_cache()
            present_trainer = pl.Trainer(max_epochs=  # args.kg_epoch,
                                         1,
                                         callbacks=[checkpoint_callback, lr_callback, present_early_stopping],
                                         accelerator="gpu", devices=1)
            train_dataloaders, val_dataloaders = model.load_kind_data(datadir=datadir, epoch=e)
            present_trainer.fit(model, train_dataloaders=train_dataloaders, val_dataloaders=val_dataloaders)
            del present_trainer, train_dataloaders, val_dataloaders
        model.switch_present_and_absent()

        for e in range(args.kg_epoch):
            torch.cuda.empty_cache()
            absent_trainer = pl.Trainer(max_epochs=  # args.kg_epoch,
                                        1,
                                        callbacks=[checkpoint_callback, lr_callback, absent_early_stopping],
                                        accelerator="gpu", devices=1)
            train_dataloaders, val_dataloaders = model.load_kind_data(datadir=datadir, epoch=e)
            absent_trainer.fit(model, train_dataloaders=train_dataloaders, val_dataloaders=val_dataloaders)
            del absent_trainer, train_dataloaders, val_dataloaders
        model.switch_present_and_absent()

        # present_trainer.fit(model, train_dataloaders=model.train_dataset['present'],
        #                     val_dataloaders=model.val_dataset['present'])
        # model.switch_present_and_absent()
        # absent_trainer.fit(model, train_dataloaders=model.train_dataset['absent'],
        #                    val_dataloaders=model.val_dataset['absent'])
        # model.switch_present_and_absent()

        print(f'save model in {model_dir}')
        torch.save(model.present_model.state_dict(), model_dir + '_present')
        torch.save(model.absent_model.state_dict(), model_dir + '_absent')
    else:
        datadir = utils.BASE_DATASET_DIR
        trainer = pl.Trainer(max_epochs=args.kg_epoch, callbacks=[checkpoint_callback, lr_callback],
                             accelerator="gpu", devices=1)

        model.load_data(datadir=datadir)
        torch.cuda.empty_cache()
        trainer.fit(model, train_dataloaders=model.train_dataset['all'],
                    val_dataloaders=model.val_dataset['all'])

        print(f'save model in {model_dir}')
        torch.save(model.model.state_dict(), model_dir)
    # trainer.save_checkpoint(model_dir)
    return model.train_present_ratio, model.train_absent_ratio, model.val_present_ratio, model.val_absent_ratio


def create_xlsx(xlsx_dir, type):
    wb = openpyxl.Workbook()
    wb.save(filename=xlsx_dir)
    wb = openpyxl.load_workbook(xlsx_dir)
    wb.create_sheet(type + '_label_word_num_map', 0)
    return wb


def store_label_word_num_map(wb, label_word_num_map, type):
    sheet = wb[type + '_label_word_num_map']
    sheet['A1'] = data_clean('all_num')
    sheet['B1'] = data_clean('present_num')
    sheet['C1'] = data_clean('present_ratio')
    sheet['D1'] = data_clean('absent_num')
    sheet['E1'] = data_clean('absent_ratio')
    for id, i in enumerate(label_word_num_map):
        sheet['A' + str(id + 2)] = data_clean(label_word_num_map[i]['all_num'])
        sheet['B' + str(id + 2)] = data_clean(label_word_num_map[i]['present_num'])
        sheet['C' + str(id + 2)] = data_clean(label_word_num_map[i]['present_ratio'])
        sheet['D' + str(id + 2)] = data_clean(label_word_num_map[i]['absent_num'])
        sheet['E' + str(id + 2)] = data_clean(label_word_num_map[i]['absent_ratio'])


def split_present_absent_labels(args, datadir, type, outputdir, two, kpdrop, kpinsert, shuffle, epoch):
    texts = read_text(os.path.join(datadir, "X." + type + ".txt"))
    indexes = read_index(os.path.join(datadir, "Y." + type + ".txt"))
    label_map = load_map(os.path.join(datadir, "output-items.txt"))
    labels_list = transfer_indexs_to_labels(label_map, indexes)  # list,需要转化成text
    overflow_ratio = 0.0
    if 'wiki' in datadir:
        for id1, label_list in enumerate(labels_list):
            for id2, label in enumerate(label_list):
                labels_list[id1][id2] = label.replace('_', ' ')
    with open(os.path.join(datadir, 'label_names_with_documents.txt'), 'w', encoding='utf-8', errors='ignore') as r:
        for id1, label_list in enumerate(labels_list):
            r.write('{labels: ')
            for id2, label in enumerate(label_list):
                r.write(label)
                r.write(',')
            r.write('\ndocument: ')
            r.write(texts[id1])
            r.write('}\n\n')
    if two:
        present_label_list, absent_label_list, present_indexes, absent_indexes, present_ratio, absent_ratio, label_word_num_map = utils.separate_present_absent_labels(
            indexes, labels_list, texts, args.match, args.stem_lambda, args.stem_delta, args.stem_model)

        xlsx_dir = os.path.join(utils.KEYPHRASE_GENERATION_RECORDS_DIR, 'res', utils.KEYPHRASE_GENERATION_MODEL_NAME,
                                type + '_intermediate_data.xlsx')
        wb = create_xlsx(xlsx_dir=xlsx_dir, type=type)
        store_label_word_num_map(wb, label_word_num_map, type)
        wb.save(xlsx_dir)

        for e in range(epoch):
            temp_present_label_list, temp_absent_label_list = present_label_list.copy(), absent_label_list.copy()
            temp_present_texts, temp_absent_texts = texts.copy(), texts.copy()
            if kpdrop is not None or kpinsert is not None:
                if kpdrop is not None:
                    temp_present_label_list, temp_absent_label_list, temp_present_texts, temp_absent_texts = utils.kpdrop(
                        present_labels=temp_present_label_list, absent_labels=temp_absent_label_list,
                        present_texts=temp_present_texts, absent_texts=temp_absent_texts,
                        texts=texts, kpdrop_type=kpdrop, kpdrop_rate=args.kpdrop_rate)
                if kpinsert is not None:
                    temp_present_label_list, temp_absent_label_list, temp_present_texts, temp_absent_texts, overflow_ratio = utils.kpinsert(
                        present_labels=temp_present_label_list, absent_labels=temp_absent_label_list,
                        present_texts=temp_present_texts, absent_texts=temp_absent_texts,
                        texts=texts, kpinsert_type=kpinsert, kpinsert_rate=args.kpinsert_rate, max_len=args.max_len)
            else:
                temp_present_texts = texts
                temp_absent_texts = texts.copy()
            if shuffle:
                utils.add_shuffle_examples(temp_present_label_list, temp_present_texts)
                utils.add_shuffle_examples(temp_absent_label_list, temp_absent_texts)
            present_labels, absent_labels = [], []
            for i in temp_present_label_list:
                present_labels.append(", ".join(i))  # 是否加prefix
            for i in temp_absent_label_list:
                absent_labels.append(", ".join(i))  # 是否加prefix

            with open(os.path.join(outputdir, "X.present." + type + "." + str(e) + ".txt"), 'w+', encoding='utf-8',
                      errors='ignore') as f:
                for i in temp_present_texts:
                    f.write(i)
                    f.write('\n')
            with open(os.path.join(outputdir, "Y_mapped.present." + type + "." + str(e) + ".txt"), 'w+',
                      encoding='utf-8',
                      errors='ignore') as f:
                for i in present_labels:
                    f.write(i)
                    f.write('\n')
            with open(os.path.join(outputdir, "X.absent." + type + "." + str(e) + ".txt"), 'w+', encoding='utf-8',
                      errors='ignore') as f:
                for i in temp_absent_texts:
                    f.write(i)
                    f.write('\n')
            with open(os.path.join(outputdir, "Y_mapped.absent." + type + "." + str(e) + ".txt"), 'w+',
                      encoding='utf-8',
                      errors='ignore') as f:
                for i in absent_labels:
                    f.write(i)
                    f.write('\n')
            with open(os.path.join(outputdir, "overflow.txt"), 'w+', encoding='utf-8',
                      errors='ignore') as f:
                f.write(str(overflow_ratio))
    else:
        labels = []
        for i in labels_list:
            labels.append(", ".join(i))  # 是否加prefix
        with open(os.path.join(outputdir, "X." + type + ".txt"), 'w+', encoding='utf-8', errors='ignore') as f:
            for i in texts:
                f.write(i)
                f.write('\n')
        with open(os.path.join(outputdir, "Y_mapped." + type + ".txt"), 'w+', encoding='utf-8',
                  errors='ignore') as f:
            for i in labels:
                f.write(i)
                f.write('\n')


def data_preprocess(args):
    datadir = utils.BASE_DATASET_DIR
    outputdir = os.path.join(utils.KEYPHRASE_GENERATION_RECORDS_DIR, 'res', utils.KEYPHRASE_GENERATION_MODEL_NAME,
                             'preprocessed data')
    if not os.path.exists(outputdir):
        os.mkdir(outputdir)
    else:
        return

    two = False
    if 'two' in args.kg_type:
        two = True

    if 'kpdrop_na' in args.kg_type:
        kpdrop = 'na'
    elif 'kpdrop_nr' in args.kg_type:
        kpdrop = 'nr'
    elif 'kpdrop_a' in args.kg_type:
        kpdrop = 'a'
    elif 'kpdrop_r' in args.kg_type:
        kpdrop = 'r'
    else:
        kpdrop = None

    if 'kpinsert_na' in args.kg_type:
        kpinsert = 'na'
    elif 'kpinsert_nr' in args.kg_type:
        kpinsert = 'nr'
    elif 'kpinsert_a' in args.kg_type:
        kpinsert = 'a'
    elif 'kpinsert_r' in args.kg_type:
        kpinsert = 'r'
    else:
        kpinsert = None

    shuffle = False
    if 'shuffle' in args.kg_type:
        shuffle = True

    split_present_absent_labels(args, datadir, 'trn', outputdir, two, kpdrop, kpinsert, shuffle, args.kg_epoch)
    split_present_absent_labels(args, datadir, 'tst', outputdir, two, kpdrop, kpinsert, shuffle, args.kg_epoch)
